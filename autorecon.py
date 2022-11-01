'''
READ THE "Instructions.txt" FILE BEFORE RUNNING

This Coding Project Serves The Following Function:
    Be able to recon all clubs thats balance at the end of the month matches the balance shown in FMS
'''
#import all the necessary libraries
import numpy as np
import pandas as pd
import os
import re
import datetime as dt
import calendar
from openpyxl import load_workbook
from openpyxl import Workbook
import xlsxwriter
from copy import copy

#file path for fms file
fms_path = "C:\\Users\\kianj\\OneDrive\\Desktop\\Python\\Work Projects\\Work Club Recon Project\\FMS FILE HERE"

#name of sheet for FMS LEDGER
fms_file = ''

#fms file pattern
fms_pattern = 'FY\d{2} - CLUB LEDGER - Period'

#file path for sports, non-sports, and off campus
sports_path = ''
non_sports_path = 'C:\\Users\\kianj\\OneDrive\\Desktop\\Python\\Work Projects\\Work Club Recon Project\\Ledgers'
off_campus_path = ''
walk_path = 'C:\\Users\\kianj\\OneDrive\\Desktop\\Python\\Work Projects\\Work Club Recon Project\\Ledgers'

#for the files that arent yet optimized
omitted_clubs = []

#create file pattern to use when finding excel sheets (anything with 9 straight digits followed by a "-")
file_pattern = r'\d{9} - '

#blank dataframe to use for ledger query
df = pd.DataFrame()

#blank dataframe for ledger recon
df_recon = pd.DataFrame()

#blank dataframe for FMS Query
df_FMS_Query = pd.DataFrame()

#blank dataframe for filtered FMS
df_FMS = pd.DataFrame()

#the actual recon sheet itself
merged_dataframes = pd.DataFrame()

#finds the FMS file name in the path alotted
def find_fms():

    #delcare fms file name as global
    global fms_file

    #create a loop to search through all excel sheets in the folder
    for folders,sub_folders,files in os.walk(fms_path):

        #searches through all files in directory
        for f in files:

            if f.endswith(".xlsx") and re.search(fms_pattern,f):
                fms_file = f

#based on FMS file name creates recon date
def find_recon_date():

    fy = int(fms_file[2:4])
    period = int(fms_file.split(" ")[6])

    if period <= 6:
        year = fy - 1 + 2000
    if period > 6:
        year = fy + 2000
    
    month = period + 6
    if month > 12:
        month = month - 12
    day = calendar.monthrange(year,month)[1]

    return f'{month}-{day}-{year}'



#create master query
def set_dataframe():

    #set the dataframe as a local variable
    global df

    #reset dataframe to empty
    df = pd.DataFrame()
    
    #create a loop to search through all excel sheets in the folder
    for folders,sub_folders,files in os.walk(walk_path):

        #searches through all files in directory
        for f in files:

            #conditional to make sure that its an excel file and that it fits the desired file pattern
            if (
                (f.endswith(".xls") or f.endswith(".xlsx"))
                and (folders == sports_path or folders == non_sports_path or off_campus_path)
                and re.search(file_pattern,f)
                and f not in omitted_clubs
                ):

                #xlrd only works with .xls and openpyxl only works with .xlsx
                if f.endswith(".xls"):
                    df_temp = pd.read_excel(folders + "\\" + f,sheet_name='Ledger',engine="xlrd")
                else:
                    df_temp = pd.read_excel(folders + "\\" + f,sheet_name='Ledger',engine="openpyxl")

                #takes the speedkey and club name from top of the sheet and makes them their own columns
                speedkey = str(df_temp.iloc[0,1])
                speedkey = speedkey[-10:]
                speedkey = int(speedkey)
                club_name = df_temp.columns[1]

                # print(club_name)
                df_temp['Speed Key'] = speedkey
                df_temp['Club Name'] = club_name

                #exception for SFA double ledger
                if(f == "9011300001 - SFA (Studnet Fashion Association).xls"):
                    df_temp.iloc[:,5] = pd.concat((df_temp.iloc[:,5],df_temp.iloc[:,11]),axis=1).sum(axis=1,min_count=1)
                    df_temp.iloc[:,7] = pd.concat((df_temp.iloc[:,7],df_temp.iloc[:,13]),axis=1).sum(axis=1,min_count=1)

                #drops the heading in the excel sheets
                df_temp = df_temp.drop([0,1,2,3],axis=0)
                cols_to_move = ['Speed Key', 'Club Name']
                df_temp = df_temp[ cols_to_move + [ col for col in df_temp.columns if col not in cols_to_move ] ]

                #drops more unecessary columns and renames columns
                df_temp = df_temp.drop(df_temp.iloc[:,10:],axis=1)
                df_temp.columns = ["Speed Key","Club Name","DATE","TYPE","FORM#","DESCRIPTION","LN REF","DEBIT","REMOVE","CREDIT"]
                df_temp = df_temp.drop("REMOVE",axis=1)

                #drops all the extra cells that have no ledger entry to record
                df_temp = df_temp.dropna(how="all",subset=['DATE', 'TYPE', 'FORM#','DESCRIPTION','DEBIT', 'CREDIT'])
                
                #creates datetime objects for the date of transaction
                df_temp['DATE'] = pd.to_datetime(df_temp['DATE']).dt.date

                #creates int64 datatypes for speedkeys
                df_temp["Speed Key"] = df_temp["Speed Key"].astype(np.int64)

                #add the temporary ledger to the df as new rows
                df = pd.concat([df,df_temp],axis=0)

    # sort from oldest to newest and reindex the dataframe
    df = df.sort_values(by="DATE")
    df.index = range(len(df.index))

#fix the speedkeys received by FMS that have the "001" or "01" Missing
def speedkey_fixer(speedkey):
    if len(str(speedkey)) == 10:
        return speedkey
    elif len(str(speedkey)) < 10:
        return speedkey * (10 ** (10 - len(str(speedkey)))) + 1
    else:
        print("Error Has Occured with speedkeys")

#create fms ledger to compare to recon ledger
def set_FMS():

    #make fms dataframe a global variable
    global df_FMS
    global df_FMS_Query

    #dataframe to use for FMS
    df_FMS = pd.read_excel(fms_path +"\\" + fms_file,
    sheet_name="Query",skiprows=1,engine="openpyxl")

    df_FMS_Query = df_FMS.copy()

    #only have data from 21601
    df_FMS = df_FMS[(df_FMS['Acct'] == 21601)]

    #only keep relevant columns
    df_FMS = df_FMS[['Component','Component Descr','Actual']]

    #sorts by speedkeys
    df_FMS = df_FMS.sort_values(by="Component")

    #reindexs the information
    df_FMS.index = range(len(df_FMS.index))

    #group all components together to get the sum of each club amount in fms
    df_FMS = df_FMS.groupby(['Component','Component Descr'],as_index=False)["Actual"].sum()

    #so we can merge tables
    df_FMS.columns = ["Speed Key","Club Name","FMS"]

    #fix the speed keys
    df_FMS["Speed Key"] = df_FMS["Speed Key"].apply(speedkey_fixer)


#create recon ledger to compare to FMS
def recon_mode(recon_date):

    #make sure that the program edits df as a global variable instead of in scope
    global df_recon

    #date of last day in ledger acceptable
    recon_date = dt.datetime.strptime(recon_date,'%m-%d-%Y')
    recon_date = recon_date.date()

    #create a copy of masterledger just for balances
    df_recon = df.copy()

    #make sure theres no transactions leading in from later transactions
    df_recon = df_recon[(df_recon["DATE"] <= recon_date)]

    # #fill null values to 0 so they can be calculated easier
    df_recon["DEBIT"] = df_recon["DEBIT"].fillna(value=0)
    df_recon["CREDIT"] = df_recon["CREDIT"].fillna(value=0)

    #create a new column to show the differences between money going in and out of ledger
    df_recon["Ledger Balance"] = df_recon["CREDIT"] - df_recon["DEBIT"]

    #find the clubs ledger balance
    df_recon = df_recon.groupby(["Speed Key","Club Name"],as_index=False)["Ledger Balance"].sum()

    #more readable
    df_recon.columns = ["Speed Key","Club Name","Ledger"]

#create method to compare the two dataframes
def compare_dataframes():

    global merged_dataframes

    merged_dataframes = df_recon.merge(df_FMS,how="inner",on="Speed Key")
    
    #get rid of FMS's Club Names
    merged_dataframes = merged_dataframes.drop("Club Name_y",axis=1)

    #fix column names
    merged_dataframes.columns = ["Speed Key","Club Name","Ledger","FMS"]

    merged_dataframes["Discrepancy"] = merged_dataframes["Ledger"] != merged_dataframes["FMS"] * -1

    merged_dataframes['Initials'] = np.nan

    clubs = []

    for index,row in merged_dataframes.iterrows():

        #if there isn't a discrepancy
        if(row.iloc[4] == False):
            merged_dataframes.at[index,"Initials"] = 'AI KA'
            
        clubs.append(row.iloc[0])
    merged_dataframes = merged_dataframes.drop("Discrepancy",axis=1)
            
    return clubs

#goes into each individual club and does their monthly recon
def ledger_recon(clubs,recon_date):

    for folders,sub_folders,files in os.walk(walk_path):

        #searches through all files in directory
        for f in files:

            #conditional to make sure that its an excel file and that it fits the desired file pattern
            if (
                (f.endswith(".xls") or f.endswith(".xlsx"))
                and (folders == sports_path or folders == non_sports_path or off_campus_path)
                and re.search(file_pattern,f)
                and int(f[0:10]) in clubs
                ):
                #open the workbook
                wb = load_workbook(folders + "\\" + f)

                #declare the two sheets in which we are copying
                ws1 = wb.worksheets[1]  #master recon sheet

                #find what the name should be for recon sheet
                month = ''
                f_year = ''
                date_list = recon_date.split("-")
                if int(date_list[0]) == 1:
                    month = "Jan"
                elif int(date_list[0]) == 2:
                    month = "Feb"
                elif int(date_list[0]) == 3:
                    month = "Mar"
                elif int(date_list[0]) == 4:
                    month = "Apr"
                elif int(date_list[0]) == 5:
                    month = "May"
                elif int(date_list[0]) == 6:
                    month = "Jun"
                elif int(date_list[0]) == 7:
                    month = "Jul"
                elif int(date_list[0]) == 8:
                    month = "Aug"
                elif int(date_list[0]) == 9:
                    month = "Sep"
                elif int(date_list[0]) == 10:
                    month = "Oct"
                elif int(date_list[0]) == 11:
                    month = "Nov"
                else:
                    month = "Dec"
                if int(date_list[0]) <= 6:
                    f_year = date_list[2]
                else:
                    f_year = int(date_list[2]) + 1

                #sheet that we are creating for this months recon
                ws2 = wb.create_sheet(month + " " + str(f_year)[2:])
                
                #copy everything from ws1 to ws2
                for row in ws1:
                    for cell in row:
                        new_cell = ws2.cell(row=cell.row, column=cell.column,
                        value= cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

                #puts the date on the recon
                ws2['A4'] = recon_date

                #get the correct row in the merged_dateframes dataframe
                ledger = ''
                fms = ''
                for index,row in merged_dataframes.iterrows():
                    if row.iloc[0] == int(f[0:10]):
                        ledger = row.iloc[2]
                        fms = row.iloc[3]

                #set the Ending Balance Per Club Ledger
                ws2['D6'] = ledger

                #set the Ending Balance per FMS
                ws2['D27'] = fms

                #Intial off and date on the recon if no discrepancy
                if ws2.cell(6,4).value + ws2.cell(27,4).value == 0:
                    ws2['B31'] = "AI KA"
                    ws2['B32'] = dt.datetime.today().strftime("%m/%d/%Y")
                
                wb.save(folders + "\\" + f)


#creates the recon sheet that we initial
def create_recon_book():
    recon_name = "(AI GENERATED) " + fms_file
    writer = pd.ExcelWriter(recon_name,engine='xlsxwriter')
    merged_dataframes.to_excel(writer,sheet_name= "Recon")
    df_FMS_Query.to_excel(writer,sheet_name= "FMS")
    df.to_excel(writer,sheet_name= "Ledger")
    writer.save()

#the run of the main program
def main():
    find_fms()
    print(f"Found FMS File! File Named: '{fms_file}'")
    recon_date = find_recon_date()
    print(f"Beginning Automated Reconcilion for {recon_date}")
    set_dataframe()
    print("Creating Ledger Recon...")
    recon_mode(recon_date)
    print("Creating FMS Recon...")
    set_FMS()
    print("Finding Discrepancies between FMS and Ledger balances...")
    clubs = compare_dataframes()
    print("Creating Recons in each Club Ledger...")
    ledger_recon(clubs,recon_date)
    print("Creating Recon File...")
    create_recon_book()
    print("File Created!")
    

if __name__ == '__main__':
    main()


'''

'''
